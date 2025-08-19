
import logging
from io import BytesIO
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404, JsonResponse
from django.contrib.auth import authenticate, login as auth_login, logout as logoutu
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import get_template
from django.utils import timezone
from django.utils.html import escape
from django.db.models import Q, DateField, DateTimeField
from xhtml2pdf import pisa
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, Image, HRFlowable, PageBreak
from reportlab.platypus.flowables import HRFlowable
from .models import Employee, SickLeave, Doctor
from collections import defaultdict
from django.contrib.auth import authenticate, login as auth_login
from django.contrib import messages
from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from html import escape
from collections import defaultdict



logger = logging.getLogger(__name__)

# --- Authentication Views ---
def base(request):
    return render(request, 'base.html')



def login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_active:
                auth_login(request, user)
                if user.is_superuser:
                    return redirect('staff')
                elif user.is_staff:
                    return redirect('master')
                else:
                    return redirect('report')  # Non-staff, non-superuser go to report
            else:
                messages.error(request, 'Your account is inactive. Contact an administrator.')
        else:
            messages.error(request, 'Invalid password or username.')
    return render(request, 'Auth/login.html')
def logout(request):
    logoutu(request)
    return redirect('login')

# --- Template Rendering Views ---
def master(request):
    return render(request, 'master.html')

def masterleave(request):
    return render(request, 'master.html')

def staff(request):
    return render(request, 'staff.html')

def developer(request):
    return render(request, 'developer.html')

def staff_list(request):
    staff_users = User.objects.filter(is_staff=True)
    return render(request, 'staff_list.html', {'staff_users': staff_users})

# --- Form Submission and Data Management ---
@login_required
def master_submit(request):
    context = {
        'employee_code': '',
        'employee_name': '',
        'department': '',
        'designation': '',
        'current_total_days': 30,
    }
    logger.debug("No doctors in context as they are fetched dynamically")

    if request.method == 'POST':
        code = request.POST.get('employee_code', '').strip()
        days_required_str = request.POST.get('days_required', '')
        start_date_str = request.POST.get('start_date', '')
        end_date_str = request.POST.get('end_date', '')
        patient_service = request.POST.get('patient_service', '')
        gender = request.POST.get('gender', '')
        doctor_remarks = request.POST.get('doctor_remarks', '')
        approved_by_id = request.POST.get('approved_by', '')
        recommendation = request.POST.get('recommendation', '')
        document_file = request.FILES.get('document_upload')
        balance_days_str = request.POST.get('balance_days', '')

        form_data = {
            'employee_code': code,
            'days_required': days_required_str,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'patient_service': patient_service,
            'gender': gender,
            'doctor_remarks': doctor_remarks,
            'approved_by': approved_by_id,
            'recommendation': recommendation,
            'balance_days': balance_days_str,
        }
        context['form_data'] = form_data

        errors = {}
        if not code:
            errors['employee_code'] = 'Employee code is required.'
        try:
            days_required = int(days_required_str)
            if days_required <= 0:
                errors['days_required'] = 'Days required must be a positive number.'
        except ValueError:
            errors['days_required'] = 'Days required must be a valid number.'
            days_required = 0

        start_date_obj = None
        if start_date_str:
            try:
                start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                errors['start_date'] = 'Invalid start date format. Use YYYY-MM-DD.'
        else:
            errors['start_date'] = 'Start date is required.'

        end_date_obj = None
        if end_date_str:
            try:
                end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                if start_date_obj and end_date_obj < start_date_obj:
                    errors['end_date'] = 'End date cannot be before the start date.'
            except ValueError:
                errors['end_date'] = 'Invalid end date format. Use YYYY-MM-DD.'
        else:
            errors['end_date'] = 'End date is required.'

        try:
            balance_days = int(balance_days_str)
        except ValueError:
            if balance_days_str:
                errors['balance_days'] = 'Balance days must be a valid number.'
            balance_days = 0

        if not approved_by_id:
            errors['approved_by'] = 'Please select a doctor for approval.'

        if errors:
            context['errors'] = errors
            context['error_message'] = 'Please correct the errors below.'
            return render(request, 'master.html', context)

        try:
            employee = Employee.objects.get(employee_code=code)
            approved_by = Doctor.objects.get(id=approved_by_id) if approved_by_id else None

            SickLeave.objects.create(
                employee=employee,
                sick_leave_days=30,
                days_required=days_required,
                start_date=start_date_obj,
                end_date=end_date_obj,
                patient_service=patient_service,
                gender=gender,
                doctor_remarks=doctor_remarks,
                recommendation=recommendation,
                approved_by=approved_by,
                document=document_file,
                created_by=request.user,
                balance_days=balance_days
            )

            success_context = {
                'message': 'Sick Leave submitted successfully.',
                'employee_code_display': employee.employee_code,
                'employee_name_display': employee.employee_name,
                'department_display': employee.department,
                'designation_display': employee.designation,
                'current_total_days': employee.current_total_days,
            }
            return render(request, 'master.html', success_context)

        except Employee.DoesNotExist:
            context['errors'] = {'employee_code': 'Employee with this code does not exist.'}
            context['error_message'] = 'Submission failed. Please check the employee code.'
            return render(request, 'master.html', context)
        except Doctor.DoesNotExist:
            context['errors'] = {'approved_by': 'Selected doctor does not exist.'}
            context['error_message'] = 'Submission failed. Please select a valid doctor.'
            return render(request, 'master.html', context)

    return render(request, 'master.html', context)

def fetch_sickleave_by_code(request):
    if request.method == 'POST':
        code = request.POST.get('employee_code')
        sickleave = None
        error = None

        if code:
            try:
                sickleave = SickLeave.objects.get(employee_code=code)
            except SickLeave.DoesNotExist:
                error = "Employee code not found."

        context = {
            'sickleave': sickleave,
            'error': error,
            'employee_code': code,
        }
        return render(request, 'staff.html', context)
    return render(request, 'staff.html')

@login_required
def upload_excel(request):
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file1')
        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect('upload_excel')

        try:
            df = pd.read_excel(excel_file)
            for _, row in df.iterrows():
                Employee.objects.update_or_create(
                    employee_code=row['employee_code'],
                    defaults={
                        'employee_name': row['employee_name'],
                        'department': row['department'],
                        'designation': row['designation'],
                    }
                )
            messages.success(request, "File uploaded and processed successfully.")
        except Exception as e:
            messages.error(request, f"Upload failed: {e}")
        return redirect('upload_excel')
    return render(request, 'developer.html')

def search_employee(request):
    context = {
        'search_type_selected': request.POST.get('search_type', 'code'),
        'search_term_input': request.POST.get('search_term', ''),
        'additional_sick_leave_input': '',
        'reason_input': '',
        'employee_details': None,
        'success_message': '',
        'error_message': ''
    }

    if request.method == 'POST':
        is_update_action = 'additional_sick_leave_days' in request.POST
        current_search_term = request.POST.get('search_term', '').strip()
        current_search_type = request.POST.get('search_type', 'code')
        context['search_term_input'] = current_search_term
        context['search_type_selected'] = current_search_type

        employee_to_display_or_update = None
        if current_search_term:
            try:
                if current_search_type == 'code':
                    employee_to_display_or_update = Employee.objects.get(employee_code=current_search_term)
                if employee_to_display_or_update:
                    context['employee_details'] = employee_to_display_or_update
                    if not is_update_action:
                        context['additional_sick_leave_input'] = employee_to_display_or_update.additional_sick_leave_days or ''
                        context['reason_input'] = employee_to_display_or_update.reason or ''
                else:
                    context['error_message'] = f"Employee with {current_search_type} '{current_search_term}' not found."
            except Employee.DoesNotExist:
                context['error_message'] = f"Employee with {current_search_type} '{current_search_term}' not found."
            except Exception as e:
                context['error_message'] = f"An error occurred: {str(e)}"
        elif not is_update_action:
            context['error_message'] = 'Please enter a search term.'

        if is_update_action and employee_to_display_or_update:
            additional_days_str = request.POST.get('additional_sick_leave_days', '').strip()
            reason_str = request.POST.get('reason', '').strip()
            context['additional_sick_leave_input'] = additional_days_str
            context['reason_input'] = reason_str

            if not additional_days_str:
                context['error_message'] = 'Please enter the number of additional sick leave days.'
            else:
                try:
                    additional_days = int(additional_days_str)
                    employee_to_display_or_update.additional_sick_leave_days = additional_days
                    employee_to_display_or_update.reason = reason_str
                    employee_to_display_or_update.current_total_days = (employee_to_display_or_update.current_total_days or 0) + additional_days
                    if employee_to_display_or_update.current_total_days < 0:
                        employee_to_display_or_update.current_total_days = 0
                    employee_to_display_or_update.save()
                    context['success_message'] = f"Successfully updated sick leave for {employee_to_display_or_update.employee_name}. New total: {employee_to_display_or_update.current_total_days} days."
                    context['employee_details'] = employee_to_display_or_update
                    context['additional_sick_leave_input'] = ''
                    context['reason_input'] = ''
                except ValueError:
                    context['error_message'] = 'Invalid number for additional sick leave days.'
                except Exception as e:
                    context['error_message'] = f'An unexpected error occurred during update: {str(e)}'
        elif is_update_action and not employee_to_display_or_update:
            context['error_message'] = 'Could not find employee to update. Please search again.'
            context['additional_sick_leave_input'] = request.POST.get('additional_sick_leave_days', '')
            context['reason_input'] = request.POST.get('reason', '')

    return render(request, 'staff.html', context)

@login_required
def fetch_employee(request):
    search_type = request.GET.get('search_type', 'code')
    search_term = request.GET.get('search_term', '').strip()

    if not search_term:
        return JsonResponse({'error': 'Enter employee code or name.'}, status=400)

    try:
        if search_type == 'code':
            employee = Employee.objects.get(employee_code__iexact=search_term)
        else:
            employee = Employee.objects.get(employee_name__iexact=search_term)
        doctors = [{'id': doctor.id, 'name': doctor.name} for doctor in Doctor.objects.all()]
        data = {
            'employee_code': employee.employee_code,
            'employee_name': employee.employee_name,
            'department': employee.department,
            'designation': employee.designation,
            'current_total_days': employee.current_total_days,
            'doctors': doctors,
        }
        return JsonResponse(data)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found.'}, status=404)



def Report(request):
    filter_type = request.GET.get('filter_type', '').strip()
    filter_value = request.GET.get('filter_value', '').strip()
    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str = request.GET.get('to_date', '').strip()
    export_to = request.GET.get('export_to', '').strip().lower()
    sickleaves_qs = SickLeave.objects.select_related('employee').all().order_by('-start_date')
    error = None
    filters_applied_text = ""

    current_year = datetime.now().year
    years_for_dropdown = [str(y) for y in range(current_year, current_year - 10, -1)]
    query_performed = bool(filter_type and (filter_value or (from_date_str and to_date_str)))

    # Annual sick leave entitlement (adjust as needed)
    ANNUAL_SICK_LEAVE_DAYS = 30  # Example: 30 days per employee per year

    if filter_type == 'date':
        if from_date_str and to_date_str:
            try:
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
                if from_date > to_date:
                    error = "'From Date' cannot be after 'To Date'."
                    sickleaves_qs = SickLeave.objects.none()
                else:
                    sickleaves_qs = sickleaves_qs.filter(start_date__lte=to_date, end_date__gte=from_date)
                    filters_applied_text = f"From {from_date_str} to {to_date_str}"
                    for leave in sickleaves_qs:
                        overlap_start = max(leave.start_date, from_date)
                        overlap_end = min(leave.end_date, to_date)
                        delta = (overlap_end - overlap_start).days + 1
                        leave.overlap_days = delta if delta > 0 else 0
            except ValueError:
                error = "Invalid date format. Use YYYY-MM-DD."
                sickleaves_qs = SickLeave.objects.none()
        else:
            error = "Please provide both From Date and To Date for date filtering."
            sickleaves_qs = SickLeave.objects.none()

    elif filter_type == 'year' and filter_value.isdigit():
        sickleaves_qs = sickleaves_qs.filter(start_date__year=int(filter_value))
        filters_applied_text = f"{filter_value}"
    elif filter_type == 'month' and filter_value.isdigit() and 1 <= int(filter_value) <= 12:
        sickleaves_qs = sickleaves_qs.filter(start_date__month=int(filter_value))
        filters_applied_text = f"{datetime(1, int(filter_value), 1).strftime('%B')}"
    elif filter_type == 'code' and filter_value:
        sickleaves_qs = sickleaves_qs.filter(employee__employee_code__icontains=filter_value)
        filters_applied_text = f"{filter_value}"
    elif filter_type == 'department' and filter_value:
        sickleaves_qs = sickleaves_qs.filter(employee__department__icontains=filter_value)
        filters_applied_text = f"'{filter_value}'"
    elif filter_type:
        error = "Invalid filter type selected."
        sickleaves_qs = SickLeave.objects.none()

    grouped = request.GET.get("grouped") == "true"
    grouped_sickleaves = []
    if not error:
        emp_leaves = defaultdict(list)
        for leave in sickleaves_qs:
            days = getattr(leave, 'overlap_days', leave.days_required)
            emp_leaves[leave.employee].append({
                'start_date': leave.start_date,
                'end_date': leave.end_date,
                'days': days,
                'pk': leave.pk,
            })

        for employee, leaves in emp_leaves.items():
            total_days = sum(l['days'] for l in leaves)
            date_ranges = ", ".join(f"{l['start_date']} to {l['end_date']}" for l in leaves)
            first_leave_pk = leaves[0]['pk'] if leaves else None
            day_balance = max(0, ANNUAL_SICK_LEAVE_DAYS - total_days)  # Prevent negative balance
            grouped_sickleaves.append({
                'employee': employee,
                'total_days': total_days,
                'day_balance': day_balance,
                'date_ranges': date_ranges,
                'first_leave_pk': first_leave_pk,
            })

    if export_to and not error:
        if export_to == 'pdf':
            return generate_sickleave_pdf_response(grouped_sickleaves, filters_applied_text, filter_type=filter_type, grouped=True)
        elif export_to == 'excel':
            return generate_sickleave_excel_response(grouped_sickleaves, filters_applied_text, grouped=True)

    context = {
        'sickleaves': grouped_sickleaves if not error and (query_performed or not filter_type) else [],
        'error': error,
        'filters': {
            'filter_type': filter_type,
            'filter_value': filter_value,
            'from_date': from_date_str,
            'to_date': to_date_str,
        },
        'years': years_for_dropdown,
        'query_performed': query_performed,
        'filters_applied': filters_applied_text,  # Pass filters_applied_text to template
    }
    return render(request, 'Auth/report.html', context)

def export_report_pdf(request):
    filter_type = request.GET.get('filter_type')
    filter_value = request.GET.get('filter_value')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    sickleaves = SickLeave.objects.all()

    if filter_type:
        if filter_type == 'year' and filter_value:
            sickleaves = sickleaves.filter(created_at__year=filter_value)
        elif filter_type == 'month' and filter_value:
            sickleaves = sickleaves.filter(created_at__month=filter_value)
        elif filter_type == 'code' and filter_value:
            sickleaves = sickleaves.filter(employee__employee_code__icontains=filter_value)
        elif filter_type == 'name' and filter_value:
            sickleaves = sickleaves.filter(employee__employee_name__icontains=filter_value)
        elif filter_type == 'date' and from_date and to_date:
            try:
                from_date_parsed = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_date_parsed = datetime.strptime(to_date, '%Y-%m-%d').date()
                sickleaves = sickleaves.filter(start_date__lte=to_date_parsed, end_date__gte=from_date_parsed)
            except ValueError:
                return HttpResponse("Invalid date format", status=400)

    template = get_template("Auth/report.html")
    html = template.render({'sickleaves': sickleaves})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sick_leave_report.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)
    return response

@login_required
def export_single_leave_pdf_view(request, leave_pk):
    try:
        if not isinstance(leave_pk, int):
            try:
                leave_pk = int(leave_pk)
            except ValueError:
                logger.error(f"export_single_leave_pdf_view received a non-integer and non-convertible leave_pk: {leave_pk}")
                return HttpResponse("Invalid Leave Record ID format.", status=400)

        leave_instance = get_object_or_404(SickLeave.objects.select_related('employee'), pk=leave_pk)
        employee = leave_instance.employee
        if not employee:
            logger.warning(f"No employee associated with SickLeave PK {leave_pk}")
            return HttpResponse("No employee associated with this leave record.", status=404)

        leave_records = SickLeave.objects.filter(employee=employee).select_related('employee').order_by('start_date')
        if not leave_records.exists():
            logger.warning(f"No sick leave records found for employee ID {employee.id}")
            return HttpResponse("No sick leave records found for this employee.", status=404)

    except Http404:
        logger.warning(f"SickLeave record with PK {leave_pk} not found.")
        return HttpResponse("Sick leave record not found. It may have been deleted.", status=404)
    except ValueError:
        logger.error(f"export_single_leave_pdf_view received an invalid value for leave_pk: {leave_pk}")
        return HttpResponse("Invalid Leave Record ID.", status=400)
    except Exception as e:
        logger.error(f"Error fetching SickLeave records for PK {leave_pk}: {e}", exc_info=True)
        return HttpResponse("An error occurred while retrieving leave records. Please contact support.", status=500)

    http_response = HttpResponse(content_type='application/pdf')
    emp_code_for_filename = escape(employee.employee_code) if employee and employee.employee_code else 'UnknownEmp'
    filename = f"SickLeave_Records_{emp_code_for_filename}.pdf"
    http_response['Content-Disposition'] = f'inline; filename="{filename}"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    style_normal = ParagraphStyle('Normal_Custom', parent=styles['Normal'], fontName='Times-Roman', fontSize=10, leading=12,
                                 alignment=TA_LEFT, spaceAfter=4, spaceBefore=4, textColor=colors.HexColor("#333333"))
    style_bold_label = ParagraphStyle('BoldLabel', parent=style_normal, fontName='Times-Bold', fontSize=10,
                                     textColor=colors.HexColor("#003087"))
    style_italic_small = ParagraphStyle('ItalicSmall', parent=styles['Normal'], fontName='Times-Italic', fontSize=8,
                                       alignment=TA_CENTER, spaceBefore=8, textColor=colors.HexColor("#666666"))
    style_company_title = ParagraphStyle('CompanyTitle', parent=styles['Heading1'], fontName='Times-Bold', fontSize=16,
                                        alignment=TA_CENTER, spaceAfter=0.15*inch, textColor=colors.HexColor("#003087"),
                                        textTransform='uppercase')
    style_report_title = ParagraphStyle('ReportTitle', parent=styles['Heading2'], fontName='Times-Bold', fontSize=12,
                                       alignment=TA_CENTER, spaceAfter=0.2*inch, textColor=colors.HexColor("#003087"))
    style_section_title = ParagraphStyle('SectionTitle', parent=styles['Heading3'], fontName='Times-Bold', fontSize=11,
                                        spaceBefore=0.2*inch, spaceAfter=0.1*inch, textColor=colors.HexColor("#003087"),
                                        alignment=TA_LEFT, backColor=colors.HexColor("#E8F0FE"))
    style_record_title = ParagraphStyle('RecordTitle', parent=styles['Normal'], fontName='Times-Bold', fontSize=11,
                                       spaceBefore=0.15*inch, spaceAfter=0.1*inch, textColor=colors.HexColor("#003087"),
                                       backColor=colors.HexColor("#E8F0FE"), leftIndent=8, rightIndent=8)

    # Header
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#003087"), spaceAfter=0.1*inch))
    elements.append(Paragraph("SUGAR CORPORATION OF UGANDA LTD LUGAZI", style_company_title))
    elements.append(Paragraph("EMPLOYEE SICK LEAVE ", style_report_title))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#003087"), spaceBefore=0.1*inch,
                              spaceAfter=0.2*inch))

    # Employee Details
    elements.append(Paragraph("Employee Information", style_section_title))
    if employee:
        elements.append(Paragraph(f"<b>Employee Code:</b> {escape(employee.employee_code or 'N/A')}", style_normal))
        elements.append(Spacer(1, 0.05*inch))
        elements.append(Paragraph(f"<b>Employee Name:</b> {escape(employee.employee_name or 'N/A')}", style_normal))
        elements.append(Spacer(1, 0.05*inch))
        emp_dept_display = escape(employee.department.name if hasattr(employee.department, 'name') else str(employee.department) or 'N/A')
        elements.append(Paragraph(f"<b>Department:</b> {emp_dept_display}", style_normal))
        if employee.additional_sick_leave_days > 0:
            elements.append(Spacer(1, 0.05*inch))
            elements.append(Paragraph(f"<b>Additional Sick Leave Days:</b> {escape(str(employee.additional_sick_leave_days or 'N/A'))}", style_normal))
            elements.append(Spacer(1, 0.05*inch))
            elements.append(Paragraph(f"<b>Reason for Additional Days:</b> {escape(employee.reason or 'N/A')}", style_normal))
    elements.append(Spacer(1, 0.2*inch))

    # Leave Details
    elements.append(HRFlowable(width="100%", thickness=1.2, color=colors.HexColor("#003087"), spaceBefore=0.1*inch,
                              spaceAfter=0.15*inch))

    model_fields = [field for field in SickLeave._meta.get_fields() if field.name not in ['employee', 'id', 'doctors_remarks', 'document', 'days_required']]
    field_names = [field.name for field in model_fields]
    field_names.insert(0, 'sick_leave_taken')

    for index, leave in enumerate(leave_records, 1):
        record_elements = []
        record_elements.append(Paragraph(f"Sick Leave {index}", style_record_title))
        record_elements.append(HRFlowable(width="70%", thickness=0.7, color=colors.HexColor("#666666"), spaceBefore=0.05*inch,
                                         spaceAfter=0.1*inch))

        table_data = []
        for field_name in field_names:
            display_name = field_name.replace('_', ' ').title()
            if field_name == 'sick_leave_taken':
                sick_leave_taken = 'N/A'
                try:
                    if leave.start_date and leave.end_date:
                        sick_leave_taken = (leave.end_date - leave.start_date).days + 1
                        sick_leave_taken = str(sick_leave_taken) if sick_leave_taken >= 0 else 'Invalid Dates'
                    else:
                        logger.warning(f"Cannot calculate sick leave taken for SickLeave PK {leave.pk}: missing date fields")
                except Exception as e:
                    logger.error(f"Error calculating sick leave taken for SickLeave PK {leave.pk}: {e}", exc_info=True)
                    sick_leave_taken = 'Error'
                value = sick_leave_taken
                display_name = "Sick Leave Taken"
            elif field_name == 'sick_leave_days':
                display_name = "Assigned Sick Leave Days"
                value = getattr(leave, field_name, None)
                value = 'N/A' if value is None else escape(str(value)).replace('\n', '<br/>')
            else:
                value = getattr(leave, field_name, None)
                if value is None:
                    value = 'N/A'
                else:
                    if isinstance(SickLeave._meta.get_field(field_name), (DateField, DateTimeField)):
                        value = value.strftime("%B %d, %Y") if value else 'N/A'
                    else:
                        value = escape(str(value)).replace('\n', '<br/>')
            table_data.append([Paragraph(f"{display_name}:", style_bold_label), Paragraph(value, style_normal)])

        record_table = Table(table_data, colWidths=[2.0*inch, 4.5*inch], hAlign='LEFT')
        table_styles = [
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
            ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor("#003087")),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E8F0FE")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#003087"))
        ]
        for i in range(len(table_data)):
            if i % 2 == 0:
                table_styles.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F9FAFB")))
            else:
                table_styles.append(('BACKGROUND', (0, i), (-1, i), colors.white))
        record_table.setStyle(TableStyle(table_styles))
        record_elements.append(record_table)
        record_elements.append(Spacer(1, 0.15*inch))

        elements.append(KeepTogether(record_elements))
        if len(leave_records) > 1 and index < len(leave_records):
            elements.append(PageBreak())

    # Footer
    elements.append(Spacer(1, 0.3 * inch))
    # current_time_str = timezone.localtime(timezone.now()).strftime("%B %d, %Y at %I:%M %p ")
    current_time_str = timezone.localtime(timezone.now()).strftime("%B %d, %Y ")
    elements.append(HRFlowable(width="100%", thickness=0.8, color=colors.HexColor("#003087"), spaceBefore=0.1*inch,
                              spaceAfter=0.1*inch))
    elements.append(Paragraph(f"<i> {current_time_str}</i>", style_italic_small))

    try:
        doc.build(elements)
        pdf_output = buffer.getvalue()
    except Exception as e:
        logger.error(f"Error building PDF for employee ID {employee.id}: {e}", exc_info=True)
        buffer.seek(0)
        buffer.truncate()
        error_doc = SimpleDocTemplate(buffer, pagesize=letter)
        error_elements = [
            Paragraph("PDF Generation Error", style_company_title),
            Paragraph(f"An error occurred while generating the PDF for employee {emp_code_for_filename}.", style_normal),
            Paragraph(f"Details: {escape(str(e))}", style_normal),
            Paragraph("Please contact support.", style_normal),
        ]
        try:
            error_doc.build(error_elements)
            pdf_output = buffer.getvalue()
        except Exception as build_e:
            logger.critical(f"FATAL: Could not build even the error PDF for employee ID {employee.id}: {build_e}", exc_info=True)
            return HttpResponse(f"A critical error occurred during PDF generation. Please contact support.", status=500, content_type="text/plain")

    buffer.close()
    http_response.write(pdf_output)
    return http_response

def generate_sickleave_pdf_response(sickleaves_data, filters_applied, filter_type='', grouped=False):
    response = HttpResponse(content_type='application/pdf')
    filename = "employee_leave_report_grouped.pdf" if grouped else "employee_leave_report_detailed.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    style_normal = ParagraphStyle('Normal_Custom', parent=styles['Normal'], fontName='Times-Roman', fontSize=10, leading=12,
                                 alignment=TA_CENTER, spaceAfter=4, spaceBefore=4, textColor=colors.HexColor("#333333"))
    style_bold_label = ParagraphStyle('BoldLabel', parent=style_normal, fontName='Times-Bold', fontSize=10,
                                     textColor=colors.HexColor("#003087"))
    style_italic_small = ParagraphStyle('ItalicSmall', parent=styles['Normal'], fontName='Times-Italic', fontSize=8,
                                       alignment=TA_CENTER, spaceBefore=8, textColor=colors.HexColor("#666666"))
    style_title = ParagraphStyle('Title', parent=styles['Heading1'], fontName='Times-Bold', fontSize=16,
                                alignment=TA_CENTER, spaceAfter=0.15*inch, textColor=colors.HexColor("#003087"),
                                textTransform='uppercase')
    style_subtitle = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontName='Times-Bold', fontSize=12,
                                   alignment=TA_CENTER, spaceAfter=0.2*inch, textColor=colors.HexColor("#003087"),
                                   textTransform='uppercase')

    elements.append(Paragraph('<hr width="100%" style="border: 1.5pt solid #003087">', style_normal))
    elements.append(Paragraph("SUGAR CORPORATION OF UGANDA LTD LUGAZI", style_title))
    
    # Updated report title to show only filters_applied
    report_title_text = f"EMPLOYEE SICK LEAVE DAYS REPORT{filters_applied and f' - {filters_applied}' or ''}"
    elements.append(Paragraph(report_title_text, style_subtitle))
    
    elements.append(Paragraph('<hr width="100%" style="border: 1pt solid #003087">', style_normal))

    data_for_table = []
    if grouped:
        headers = ["No", "Emp Code", "Name", "Department", "Days Taken", "Balance days"]
        data_for_table.append(headers)
        col_widths = [0.4*inch, 0.8*inch, 2.0*inch, 1.5*inch, 1.2*inch, 1.2*inch, 2.5*inch]

        for i, record in enumerate(sickleaves_data, 1):
            employee_data = record.get('employee')
            sick_leave_taken = record.get('total_days', 0)
            day_balance = record.get('day_balance', 0)
            date_ranges_str = record.get('date_ranges', '')

            emp_code = employee_data.employee_code or 'N/A' if hasattr(employee_data, 'employee_code') else 'N/A'
            emp_name = employee_data.employee_name or 'N/A' if hasattr(employee_data, 'employee_name') else 'N/A'
            department = employee_data.department or 'N/A' if hasattr(employee_data, 'department') else 'N/A'

            row_data = [
                str(i),
                Paragraph(escape(emp_code), style_normal),
                Paragraph(escape(emp_name), style_normal),
                Paragraph(escape(str(department)), style_normal),
                str(sick_leave_taken),
                str(day_balance),
                
            ]
            data_for_table.append(row_data)
    else:
        headers = ["No", "Emp Code", "Name", "Department", "Assigned Sick Leave Days", "From Date", "To Date"]
        data_for_table.append(headers)
        col_widths = [0.4*inch, 0.8*inch, 2.0*inch, 1.8*inch, 1.2*inch, 1.0*inch, 1.0*inch]

        for i, leave in enumerate(sickleaves_data, 1):
            row_data = [
                str(i),
                Paragraph(escape(leave.employee.employee_code or 'N/A'), style_normal),
                Paragraph(escape(leave.employee.employee_name or 'N/A'), style_normal),
                Paragraph(escape(str(leave.employee.department or 'N/A')), style_normal),
                str(leave.sick_leave_days or 'N/A'),
                leave.start_date.strftime("%Y-%m-%d") if leave.start_date else 'N/A',
                leave.end_date.strftime("%Y-%m-%d") if leave.end_date else 'N/A'
            ]
            data_for_table.append(row_data)

    if len(data_for_table) > 1:
        table = Table(data_for_table, colWidths=col_widths, repeatRows=1)
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E8F0FE")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#003087")),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor("#333333")),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
            ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor("#003087")),
            ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (1, 1), (3, -1), 'LEFT'),
            ('ALIGN', (6, 1), (6, -1), 'LEFT') if grouped else ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ]
        for i in range(1, len(data_for_table)):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F9FAFB")))
        table.setStyle(TableStyle(table_style))
        elements.append(KeepTogether(table))
    else:
        elements.append(Paragraph("No data found for the selected criteria.", style_normal))
        elements.append(Spacer(1, 0.1*inch))

    elements.append(Spacer(1, 0.3*inch))
    current_time_str = timezone.localtime(timezone.now()).strftime("%B %d, %Y")
    elements.append(Paragraph('<hr width="100%" style="border: 0.8pt solid #003087">', style_normal))
    elements.append(Paragraph(f"<i>{current_time_str}</i>", style_italic_small))

    try:
        doc.build(elements)
        pdf = buffer.getvalue()
    except Exception as e:
        buffer.seek(0)
        buffer.truncate()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = [
            Paragraph("PDF Generation Error", style_title),
            Paragraph(f"An error occurred while generating the PDF: {escape(str(e))}", style_normal),
            Paragraph("Please contact support.", style_normal)
        ]
        doc.build(elements)
        pdf = buffer.getvalue()

    buffer.close()
    response.write(pdf)
    return response



def generate_sickleave_excel_response(sickleaves_data, filters_applied, grouped=False):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "employee_leave_report_grouped.xlsx" if grouped else "employee_leave_report_detailed.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leave Report"

    # Title
    sheet.merge_cells('A1:G1')
    title_cell = sheet['A1']
    title_cell.value = f"EMPLOYEE SICK LEAVE DAYS REPORT{filters_applied and f' - {filters_applied}' or ''}"
    title_cell.font = Font(name='Times New Roman', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    current_row = 2
    headers = ["No", "Employee Code", "Name", "Department", "Days Taken", "Balance days",] if grouped else ["No", "Employee Code", "Name", "Department", "Days Required", "From", "To"]
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=current_row, column=col_num, value=header_title)
        cell.font = Font(name='Times New Roman', bold=True)
        cell.alignment = Alignment(horizontal='center')

    data_row_start = current_row + 1
    if grouped:
        for i, item in enumerate(sickleaves_data, 1):
            current_row += 1
            emp = item['employee']
            row_data = [
                i,
                emp.employee_code or 'N/A',
                emp.employee_name or 'N/A',
                emp.department or 'N/A',
                item['total_days'],
                item.get('day_balance', 'N/A'),
                
            ]
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col_num, value=cell_value)
                cell.font = Font(name='Times New Roman')
                cell.alignment = Alignment(horizontal='center')
    else:
        for i, leave in enumerate(sickleaves_data, 1):
            current_row += 1
            row_data = [
                i,
                leave.employee.employee_code or 'N/A',
                leave.employee.employee_name or 'N/A',
                leave.employee.department or 'N/A',
                leave.days_required or 'N/A',
                leave.start_date,
                leave.end_date
            ]
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col_num, value=cell_value)
                cell.font = Font(name='Times New Roman')
                cell.alignment = Alignment(horizontal='center')
                if isinstance(cell_value, datetime):
                    cell.number_format = 'YYYY-MM-DD'

    # Adjust column widths
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row_idx in range(data_row_start, current_row + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val:
                max_length = max(max_length, len(str(val)))
        sheet.column_dimensions[column_letter].width = max_length + 2

    # No data message
    if not sickleaves_data:
        sheet.cell(row=current_row + 1, column=1, value="No data found for the selected criteria.").font = Font(name='Times New Roman')

    # Footer with current date
    current_row += 2
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    footer_cell = sheet.cell(row=current_row, column=1)
    footer_cell.value = timezone.localtime(timezone.now()).strftime("%B %d, %Y")
    footer_cell.font = Font(name='Times New Roman', italic=True, size=8)
    footer_cell.alignment = Alignment(horizontal='right')

    workbook.save(response)
    return response